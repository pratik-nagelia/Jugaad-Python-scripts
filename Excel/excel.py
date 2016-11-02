import openpyxl
wb = openpyxl.load_workbook('Counts.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
sheett = wb.get_sheet_by_name('Sheet2')
for i in range(1,13):
	for j in range(1,7):
		sheett.cell(row = j ,column = i).value = 0
for i in range(2,212):
	x = int(sheet.cell(row = i,column =2).value) - 2010
	y = sheet.cell(row = i,column =1).value
	z = sheet.cell(row = i,column =3).value
	val = sheett.cell(row = x, column = y).value
	sheett.cell(row = x, column = y).value = val + z

wb.save('pythonRocks.xlsx')